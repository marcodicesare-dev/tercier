#!/usr/bin/env python3
"""
Lumina AG — 48-Month Financial Model (FULLY ITEMIZED, RESEARCHED)
Every single cost on its own row. Zug + Zurich comparison. April 5 2026.
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = "lumina-financial-model-v2.xlsx"
M = 48
EUR_CHF, USD_CHF = 0.9224, 0.80
def e2c(e): return round(e * EUR_CHF)
def u2c(u): return round(u * USD_CHF)
def c2e(c): return c / EUR_CHF
CAPITAL = 200_000

# ARPU
CA = {1:1000,2:1400,3:1800}; IA = {1:1000,2:1500,3:2000}
def phase(m):
    if m>=37: return 3
    elif m>=24: return 2
    return 1

# Chains & Indies
CHAINS=[("Kempinski",1,82),("Radisson",12,95),("Barceló",16,70),("Mandarin Oriental",20,36),
        ("Hyatt Luxury",24,45),("Jumeirah",28,26),("Meliá",32,80),("Rosewood",36,30),
        ("Minor Hotels",40,55),("Rotana",44,42)]
def ca(i,m):
    _,s,t=CHAINS[i]
    if m<s: return 0
    e=m-s
    if e>=23: return round(t*.55)
    elif e>=17: return round(t*.50)
    elif e>=11: return round(t*.40)
    elif e>=5: return round(t*.15)
    return 3
def cp(i,m):
    _,s,_=CHAINS[i]; return ca(i,m) if m>=s+3 else 0
def indie():
    c=[0]*49
    for m in range(1,49):
        if m<7: c[m]=0
        elif m==7: c[m]=2
        else:
            p=c[m-1]; yr=(m-7)//12; c[m]=max(0,p+round(2*(1.20)**yr)-round(p*.12/12))
    return c
IND = indie()

# CEO (researched social contributions per level)
CEO_COST = {106080: 9737, 150000: 13689, 185000: 15893, 220000: 19945}
# Interpolate 185K: gross 15417 + social ~(185000*0.0951/12+15417*0... let me compute properly
# 185K: AHV/IV/EO 5.3% = 9805, ALV 1.1% on 148200 = 1630, ALV sol 0.5% on 36800 = 184,
#   UVG 0.17% on 148200 = 252, BVG 3.5% on 64260 = 2249, FAK 1.35% = 2498, Admin 0.1% = 185
#   Total social = 16803/yr = 1400/mo. Gross = 15417/mo. Total = 16817/mo
# Actually let me just compute: 185000/12 = 15417 gross. Social ~9.5% = 14600/yr = 1217/mo. Total = 16634
# Hmm let me be precise:
# AHV/IV/EO 5.3% × 185000 = 9805
# ALV 1.1% × 148200 = 1630.20
# ALV sol 0.5% × (185000-148200) = 0.5% × 36800 = 184
# UVG 0.17% × 148200 = 251.94
# BVG 3.5% × 64260 = 2249.10
# FAK 1.35% × 185000 = 2497.50
# Admin 0.1% × 185000 = 185
# Total social = 16802.74/yr = 1400.23/mo
# Gross/mo = 15416.67
# Total/mo = 16816.90 ≈ 16817
CEO_COST[185000] = 16817
CEO_STEPUPS = [(300000,150000),(1500000,185000),(3000000,220000)]

# ─── EVERY SINGLE COST ITEM ───
# Format: (label, values_for_48_months_as_list)
# We'll build these programmatically then write each as its own row

def tier(m):
    if m>=37: return 3
    elif m>=25: return 2
    elif m>=13: return 1
    return 0

def build_monthly():
    """Pre-compute ALL data for each month."""
    data = []
    cash = CAPITAL; prev_out = 0
    for m in range(1, 49):
        d = {}
        p = phase(m); t = tier(m)
        # Hotels
        d["chain_active"] = [ca(i,m) for i in range(10)]
        d["total_chain_active"] = sum(d["chain_active"])
        d["chain_pay"] = [cp(i,m) for i in range(10)]
        d["total_chain_pay"] = sum(d["chain_pay"])
        d["indie"] = IND[m]
        d["total_active"] = d["total_chain_active"] + d["indie"]
        d["total_pay"] = d["total_chain_pay"] + d["indie"]
        # Revenue
        d["rev_chain"] = e2c(d["total_chain_pay"] * CA[p])
        d["rev_indie"] = e2c(d["indie"] * IA[p])
        d["rev"] = d["rev_chain"] + d["rev_indie"]
        d["arr"] = d["rev"] * 12
        d["arr_eur"] = c2e(d["arr"])
        # CEO
        sal = 106080
        for th, s in CEO_STEPUPS:
            # Runway clause: cash >= net_burn × 6. If profitable (net_burn <= 0), runway = infinite.
            prev_net_burn = prev_out - (data[-1]["rev"] if data else 0)
            clause_met = (m == 1) or (prev_net_burn <= 0) or (cash >= prev_net_burn * 6)
            if d["arr_eur"] >= th and clause_met: sal = s
            else: break
        d["ceo_sal"] = sal
        d["ceo_gross"] = round(sal/12)
        d["ceo_social"] = CEO_COST[sal] - d["ceo_gross"]
        d["ceo_total"] = CEO_COST[sal]
        # Team members individually
        team_def = [("dev",4000,1),("pm",4000,1),("cs",5000,12),("seneng",7000,18),
                    ("sales",6000,18),("eng",6000,30),("mktg",5000,36)]
        for key, eur, start in team_def:
            d[f"team_{key}"] = e2c(eur) if m >= start else 0
        d["team_total"] = sum(d[f"team_{k}"] for k,_,_ in team_def)
        # AI COGS
        d["ai_cogs"] = e2c(d["total_pay"] * 50)
        # Dev tooling (3 items, USD)
        d["tool_claude"] = u2c(200)
        d["tool_codex"] = u2c(200)
        d["tool_perplexity"] = u2c(200)
        # Infrastructure (individual SaaS, USD, tiered)
        infra = {
            # (label, M1-12, M13-24, M25-36, M37-48) in USD — REAL usage costs
            "supabase": [75,110,150,200],
            "vercel": [150,200,300,400],
            "dataforseo": [75,150,200,250],
            "serpapi": [75,75,150,150],
            "firecrawl": [47,47,47,47],
            "google_ws": [35,50,70,70],
            "domain_renew": [7,7,7,7],
            "onepassword": [24,40,56,64],
            "canva": [13,13,13,13],
            "sentry": [26,26,26,26],
            "resend": [20,20,20,20],
            "notion": [20,30,30,30],
            "cloudflare": [20,20,20,20],
            "slack": [25,40,56,64],
            "github_actions": [15,20,30,40],
            "upstash_redis": [10,20,30,30],
            "posthog": [0,25,50,50],
            "misc_overages": [25,30,40,50],
        }
        for key, tiers in infra.items():
            d[f"infra_{key}"] = u2c(tiers[t])
        # Admin (CHF native, tiered)
        admin = {
            "treuhand": [300,300,500,700],
            "legal": [250,250,300,400],
            "d_o": [167,167,200,250],
            "cyber": [34,34,50,75],
            "liability": [100,100,100,100],
            "banking": [30,30,50,50],
        }
        for key, tiers in admin.items():
            d[f"admin_{key}"] = tiers[t]
        # Workspace
        d["ws_internet"] = 40
        d["ws_coworking"] = [0, 350, 500, 700][t]
        # S&M
        d["sm_travel"] = [400,1000,1500,2000][t]
        d["sm_marketing"] = [500,1500,2500,3500][t]
        # Processing
        d["processing"] = round(d["rev"] * 0.03)
        # Contingency
        d["contingency"] = 300
        # One-time (M1 only, each item)
        ot_items = {
            "ot_nexova": 690, "ot_register": 650, "ot_notary_zug": 1000,
            "ot_capital_dep": 250, "ot_signatures": 100,
            "ot_macbook": 3700, "ot_display": 1290, "ot_keyboard": 83,
            "ot_mouse": 69, "ot_monitor_arm": 165,
            "ot_legal": 4000, "ot_domain": 128,
        }
        for key, val in ot_items.items():
            d[key] = val if m == 1 else 0
        # Collect ALL cost keys (in display order)
        d["_cost_keys"] = (
            ["ceo_gross", "ceo_social"]
            + [f"team_{k}" for k,_,_ in team_def]
            + ["ai_cogs", "tool_claude", "tool_codex", "tool_perplexity"]
            + [f"infra_{k}" for k in infra]
            + [f"admin_{k}" for k in admin]
            + ["ws_internet", "ws_coworking", "sm_travel", "sm_marketing",
               "processing", "contingency"]
            + list(ot_items.keys())
        )
        d["total_costs"] = sum(d[k] for k in d["_cost_keys"])
        d["ebitda"] = d["rev"] - d["total_costs"]
        # Cash
        d["cash_in"] = (CAPITAL if m==1 else 0) + d["rev"]
        d["closing"] = (CAPITAL + d["rev"] - d["total_costs"]) if m==1 else (cash + d["rev"] - d["total_costs"])
        # Runway = cash / net_burn. If profitable, runway = infinite (999).
        prev_net_burn = prev_out - (data[-1]["rev"] if data else 0)
        if prev_net_burn <= 0:
            d["runway"] = 999  # profitable = infinite runway
            d["clause"] = "YES"
        else:
            d["runway"] = cash / prev_net_burn if prev_net_burn > 0 else 999
            d["clause"] = "YES" if (m == 1 or cash >= prev_net_burn * 6) else "NO"
        prev_out = d["total_costs"]; cash = d["closing"]
        data.append(d)
    return data

# Labels for every cost row
COST_LABELS = {
    "ceo_gross": "CEO gross salary (CHF)",
    "ceo_social": "CEO employer social (10.1%)",
    "team_dev": "Developer FT (€4K/mo)",
    "team_pm": "PM PT (€4K/mo)",
    "team_cs": "CS/Implementation (€5K/mo, M12+)",
    "team_seneng": "Senior Engineer (€7K/mo, M18+)",
    "team_sales": "Sales/BD (€6K/mo, M18+)",
    "team_eng": "Engineer (€6K/mo, M30+)",
    "team_mktg": "Marketing (€5K/mo, M36+)",
    "ai_cogs": "AI COGS (€50/hotel/mo)",
    "tool_claude": "Claude Code Max 20x ($200)",
    "tool_codex": "OpenAI Codex Pro ($200)",
    "tool_perplexity": "Perplexity Max ($200)",
    "infra_supabase": "Supabase Pro + compute",
    "infra_vercel": "Vercel Pro",
    "infra_dataforseo": "DataForSEO usage",
    "infra_serpapi": "SerpApi Developer",
    "infra_firecrawl": "Firecrawl",
    "infra_google_ws": "Google Workspace",
    "infra_domain_renew": "Domain .ai renewal",
    "infra_onepassword": "1Password Business",
    "infra_canva": "Canva Pro",
    "infra_sentry": "Sentry monitoring",
    "infra_resend": "Resend email",
    "infra_notion": "Notion",
    "infra_cloudflare": "Cloudflare Pro",
    "infra_slack": "Slack Pro",
    "infra_github_actions": "GitHub Actions (CI/CD)",
    "infra_upstash_redis": "Upstash Redis (caching)",
    "infra_posthog": "PostHog analytics",
    "infra_misc_overages": "Misc SaaS + overages",
    "admin_treuhand": "Treuhand / accounting",
    "admin_legal": "Legal retainer",
    "admin_d_o": "D&O insurance (CHF 1M cover)",
    "admin_cyber": "Cyber insurance",
    "admin_liability": "General liability insurance",
    "admin_banking": "Banking + FX (UBS/Wise)",
    "ws_internet": "Internet (Sunrise 1Gbit)",
    "ws_coworking": "Coworking (Westhive Zug)",
    "sm_travel": "Travel (client visits)",
    "sm_marketing": "Marketing & events",
    "processing": "Payment processing (Stripe 3%)",
    "contingency": "Contingency buffer",
    "ot_nexova": "One-time: Nexova AG Plus formation",
    "ot_register": "One-time: Commercial register entry",
    "ot_notary_zug": "One-time: Notary fee (Zug tariff)",
    "ot_capital_dep": "One-time: Capital deposit account",
    "ot_signatures": "One-time: Signature certifications",
    "ot_macbook": "One-time: MacBook Pro 16\" M5 Pro 64GB/2TB",
    "ot_display": "One-time: Apple Studio Display 2026",
    "ot_keyboard": "One-time: Logitech MX Keys S",
    "ot_mouse": "One-time: Logitech MX Master 3S",
    "ot_monitor_arm": "One-time: Ergotron LX monitor arm",
    "ot_legal": "One-time: Legal (SHA, PSOP, contracts)",
    "ot_domain": "One-time: Domain .ai (2yr reg)",
}

# ─── FORMATTING ───
INK="1A120B"; CREAM="F5EFE6"; TERRA="C17F59"; DEEP="8B4A2B"; GOLD="C9A96E"; W="FFFFFF"
FT=Font(name="Calibri",size=14,bold=True,color=INK)
FH=Font(name="Calibri",size=10,bold=True,color=W)
FS=Font(name="Calibri",size=10,bold=True,color=INK)
FN=Font(name="Calibri",size=10,color=INK)
FI=Font(name="Calibri",size=10,color="000088")
FF=Font(name="Calibri",size=10,color=INK)
FM=Font(name="Calibri",size=9,italic=True,color="888888")
PH=PatternFill(start_color=DEEP,end_color=DEEP,fill_type="solid")
PS=PatternFill(start_color=CREAM,end_color=CREAM,fill_type="solid")
AC=Alignment(horizontal="center",vertical="center",wrap_text=True)
BT=Border(bottom=Side(style="thin",color=GOLD))
BB=Border(top=Side(style="medium",color=DEEP),bottom=Side(style="double",color=DEEP))
FC='"CHF "#,##0;[Red]"CHF "\\(#,##0\\)'; FE='#,##0'; FP='0.0%'; FN_='#,##0'
def col(m): return get_column_letter(m+1)
def mc(m): return m+1
def yc(y): return get_column_letter(M+1+y)

def build():
    data = build_monthly()
    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 1: ASSUMPTIONS (master reference)
    # ═══════════════════════════════════════════════════════════════════
    ws = wb.active; ws.title = "Assumptions"; ws.sheet_properties.tabColor = TERRA
    ws.column_dimensions["A"].width = 55
    for c in "BCDEFGHI": ws.column_dimensions[c].width = 18
    r = [1]  # mutable counter

    def w(lbl, val=None, fmt=None, note=None, font=FN):
        ws.cell(r[0],1,lbl).font = font
        if val is not None: ws.cell(r[0],2,val).font = FI
        if fmt: ws.cell(r[0],2).number_format = fmt
        if note: ws.cell(r[0],3,note).font = FM
        r[0] += 1
    def sec(txt): ws.cell(r[0],1,txt).font = FS; ws.cell(r[0],1).fill = PS; r[0] += 1
    def blk(): r[0] += 1

    ws.cell(1,1,"LUMINA — ALL ASSUMPTIONS (Researched April 2026)").font = FT; r[0] = 3
    sec("FX RATES")
    w("EUR/CHF", EUR_CHF, "0.0000", "1 EUR = 0.9224 CHF. Multiply. ECB Apr 3 2026")
    w("USD/CHF", USD_CHF, "0.0000", "1 USD = 0.80 CHF. Multiply. Apr 4 2026")
    w("€1,000 → CHF", e2c(1000), FC); w("$600 → CHF", u2c(600), FC); blk()
    sec("CAPITAL"); w("Initial capital (CHF)", CAPITAL, FC, "Swiss AG share capital, hard CHF"); blk()
    sec("ARPU (EUR/hotel/month)")
    for ci,h in enumerate(["","Phase 1 M1-23","Phase 2 M24-36","Phase 3 M37-48"],1):
        ws.cell(r[0],ci,h).font = FH if ci>1 else FN
        if ci>1: ws.cell(r[0],ci).fill = PH; ws.cell(r[0],ci).alignment = AC
    r[0]+=1
    ws.cell(r[0],1,"Chain").font=FN
    for i,p in enumerate([1,2,3],2): ws.cell(r[0],i,CA[p]).font=FI
    r[0]+=1
    ws.cell(r[0],1,"Indie").font=FN
    for i,p in enumerate([1,2,3],2): ws.cell(r[0],i,IA[p]).font=FI
    r[0]+=1; blk()
    sec("CHAIN SCHEDULE (10 chains, 48 months)")
    for nm,s,t in CHAINS: w(f"  {nm}: M{s}, {t} hotels total")
    w("  Trial: 3 hotels free for 3 months. Rollout pays immediately.", font=FM); blk()
    sec("INDEPENDENT HOTELS")
    w("Start M7, 2/mo base, +20% YoY growth, 12% annual churn"); blk()
    sec("CEO SALARY (CHF, researched social contributions)")
    w("CHF 106,080 → company cost CHF 9,737/mo", 9737, FC, "10.14% employer social")
    w("CHF 150,000 → company cost CHF 13,689/mo", 13689, FC, "9.51%")
    w("CHF 185,000 → company cost CHF 16,817/mo", 16817, FC, "9.09%")
    w("CHF 220,000 → company cost CHF 19,945/mo", 19945, FC, "8.79%")
    w("Step-ups: €300K→150K, €1.5M→185K, €3M→220K + runway clause", font=FM); blk()
    sec("TEAM (EUR/mo total company cost per person)")
    for role,eur,start in [("Developer FT",4000,1),("PM PT",4000,1),("CS",5000,12),
        ("Senior Eng",7000,18),("Sales/BD",6000,18),("Engineer",6000,30),("Marketing",5000,36)]:
        w(f"  {role}: €{eur:,} = CHF {e2c(eur):,}/mo from M{start}")
    blk()
    sec("DEV TOOLING ($600 USD = CHF 480)")
    w("  Claude Code Max 20x", u2c(200), FC, "$200"); w("  OpenAI Codex Pro", u2c(200), FC, "$200"); w("  Perplexity Max", u2c(200), FC, "$200"); blk()
    sec("INFRASTRUCTURE (each SaaS tool, USD → CHF)")
    for nm,costs in [("Supabase Pro+compute",[75,110,150,200]),("Vercel Pro+usage",[150,200,300,400]),
        ("DataForSEO",[75,150,200,250]),("SerpApi",[75,75,150,150]),("Firecrawl Standard",[47,47,47,47]),
        ("Google Workspace",[35,50,70,70]),("Domain renewal",[7,7,7,7]),("1Password",[24,40,56,64]),
        ("Canva Pro",[13,13,13,13]),("Sentry Team",[26,26,26,26]),("Resend Pro",[20,20,20,20]),
        ("Notion Plus",[20,30,30,30]),("Cloudflare Pro",[20,20,20,20]),("Slack Pro",[25,40,56,64]),
        ("GitHub Actions",[15,20,30,40]),("Upstash Redis",[10,20,30,30]),
        ("PostHog",[0,25,50,50]),("Misc+overages",[25,30,40,50])]:
        ws.cell(r[0],1,f"  {nm}").font=FN
        for ci,v in enumerate(costs,2): ws.cell(r[0],ci,f"${v}→CHF {u2c(v)}").font=FM
        r[0]+=1
    blk()
    sec("ADMIN (CHF native)")
    for nm,costs in [("Treuhand",[300,300,500,700]),("Legal",[250,250,300,400]),
        ("D&O ins.",[167,167,200,250]),("Cyber ins.",[34,34,50,75]),
        ("Liability",[100,100,100,100]),("Banking",[30,30,50,50])]:
        ws.cell(r[0],1,f"  {nm}").font=FN
        for ci,v in enumerate(costs,2): ws.cell(r[0],ci,v).font=FI; ws.cell(r[0],ci).number_format=FC
        r[0]+=1
    blk()
    sec("WORKSPACE (CHF)")
    w("  Internet: CHF 40/mo (Sunrise 1Gbit)"); w("  Coworking: M1-12 CHF 0, M13 CHF 350, M25 CHF 500, M37 CHF 700"); blk()
    sec("ONE-TIME COSTS (CHF, M1)")
    for nm, val in [("Nexova AG Plus formation",690),("Commercial register",650),("Notary Zug",1000),
        ("Capital deposit account",250),("Signature certs ×5",100),("MacBook Pro 16\" M5 Pro 64GB/2TB",3700),
        ("Apple Studio Display 2026",1290),("Logitech MX Keys S keyboard",83),("Logitech MX Master 3S mouse",69),
        ("Ergotron LX monitor arm",165),("Legal (SHA, PSOP, contracts)",4000),("Domain .ai (2yr)",128)]:
        w(f"  {nm}", val, FC)
    w("  TOTAL ONE-TIME", sum([690,650,1000,250,100,3700,1290,83,69,165,4000,128]), FC, font=FS); blk()
    sec("EQUITY"); w("Amedeo 40% / Corsaro 40% / Marco 20% base / PSOP 5% off cap")
    sec("EXIT KICKER"); w("Under €30M: 20% | €30M+: 23% | €50M+: 25% | €100M+ (cap): 30%")
    sec("VESTING"); w("M12: +5% (15+ paying) | M24: +7% (ARR≥€500K) | M36: +8% (ARR≥€1.5M)")
    sec("ANTI-DILUTION"); w("10% floor through Series A, 8% from Series B. Pro-rata."); blk()
    sec("ZUG vs ZURICH"); w("Baar/Zug profit tax: 11.82%",None,None,"Saves CHF 38K-78K/yr vs Zurich at 19.59%")
    w("Zurich City profit tax: 19.59%",None,None,"+7.77pp higher")
    w("Zurich notary: +CHF 500",None,None,"CHF 1,500 vs CHF 1,000")
    w("FAK Zurich: 0.98% vs Zug 1.35%",None,None,"CHF 400/yr cheaper per employee")

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 2: MONTHLY P&L — EVERY item on its own row
    # ═══════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Monthly P&L"); ws2.sheet_properties.tabColor = DEEP
    ws2.column_dimensions["A"].width = 44
    for m in range(1,49): ws2.column_dimensions[col(m)].width = 13
    for y in range(1,5): ws2.column_dimensions[yc(y)].width = 16
    ws2.freeze_panes = "B5"

    r2 = 1; ws2.cell(r2,1,"LUMINA — MONTHLY P&L (CHF) — ZUG/BAAR").font = FT; r2 += 2
    for m in range(1,49): yr=(m-1)//12+1; ws2.cell(r2,mc(m),f"Y{yr}").font=Font(name="Calibri",size=9,color=TERRA); ws2.cell(r2,mc(m)).alignment=AC
    for y in range(1,5): ws2.cell(r2,M+1+y,f"Y{y}").font=FH; ws2.cell(r2,M+1+y).fill=PH; ws2.cell(r2,M+1+y).alignment=AC
    r2+=1
    ws2.cell(r2,1,"Month").font=FH; ws2.cell(r2,1).fill=PH
    for m in range(1,49): ws2.cell(r2,mc(m),f"M{m}").font=FH; ws2.cell(r2,mc(m)).fill=PH; ws2.cell(r2,mc(m)).alignment=AC
    r2+=1

    def wr2(key, lbl, fmt=FC, bold=False, bdr=None):
        nonlocal r2; f=FS if bold else FF
        ws2.cell(r2,1,lbl).font=f
        if bdr: ws2.cell(r2,1).border=bdr
        for m in range(1,49):
            ws2.cell(r2,mc(m),data[m-1][key]).font=f; ws2.cell(r2,mc(m)).number_format=fmt
            if bdr: ws2.cell(r2,mc(m)).border=bdr
        for y in range(1,5):
            s,e=(y-1)*12,y*12
            v = data[e-1][key] if fmt==FN_ else sum(data[i][key] for i in range(s,e))
            ws2.cell(r2,M+1+y,v).font=f; ws2.cell(r2,M+1+y).number_format=fmt
        rr=r2; r2+=1; return rr

    # HOTEL COUNTS
    ws2.cell(r2,1,"HOTEL COUNTS").font=FS; ws2.cell(r2,1).fill=PS; r2+=1
    for idx,(nm,_,_) in enumerate(CHAINS):
        ws2.cell(r2,1,f"  {nm}").font=FN
        for m in range(1,49): ws2.cell(r2,mc(m),data[m-1]["chain_active"][idx]).font=FF; ws2.cell(r2,mc(m)).number_format=FN_
        for y in range(1,5): ws2.cell(r2,M+1+y,data[y*12-1]["chain_active"][idx]).font=FF; ws2.cell(r2,M+1+y).number_format=FN_
        r2+=1
    wr2("total_chain_active","Total chain active",FN_,True,BT)
    wr2("indie","Independent active",FN_)
    wr2("total_active","TOTAL ACTIVE",FN_,True,BT)
    wr2("total_chain_pay","Chain paying (post-trial)",FN_)
    wr2("indie","Indie paying",FN_)
    r_tp = wr2("total_pay","TOTAL PAYING",FN_,True,BB)
    r2+=1

    # REVENUE
    ws2.cell(r2,1,"REVENUE (CHF)").font=FS; ws2.cell(r2,1).fill=PS; r2+=1
    wr2("rev_chain","Chain revenue")
    wr2("rev_indie","Indie revenue")
    r_rev = wr2("rev","TOTAL REVENUE",FC,True,BB)

    ws2.cell(r2,1,"ARR (CHF)").font=FS
    for m in range(1,49): c=col(m); ws2.cell(r2,mc(m),f"={c}{r_rev}*12").font=FS; ws2.cell(r2,mc(m)).number_format=FC
    for y in range(1,5): ec=col(y*12); ws2.cell(r2,M+1+y,f"={ec}{r2}").font=FS; ws2.cell(r2,M+1+y).number_format=FC
    r_arr=r2; r2+=1

    ws2.cell(r2,1,"ARR (EUR)").font=FM
    for m in range(1,49): c=col(m); ws2.cell(r2,mc(m),f"={c}{r_arr}/{EUR_CHF}").font=FM; ws2.cell(r2,mc(m)).number_format=FE
    for y in range(1,5): ec=col(y*12); ws2.cell(r2,M+1+y,f"={ec}{r2}").font=FM; ws2.cell(r2,M+1+y).number_format=FE
    r2+=2

    # COSTS — EVERY ITEM
    ws2.cell(r2,1,"COSTS (CHF) — EVERY ITEM").font=FS; ws2.cell(r2,1).fill=PS; r2+=1

    # Write section labels between groups
    cost_keys = data[0]["_cost_keys"]
    cost_rows = {}  # key → row number
    sections = [
        (0, "People"),
        (2, None),  # team starts
        (9, "AI & Technology"),
        (13, "Infrastructure (SaaS)"),
        (26, "Admin & Professional"),
        (32, "Workspace"),
        (34, "Sales & Marketing"),
        (36, "Variable & Other"),
        (38, "One-Time (M1 only)"),
    ]
    sec_idx = {s[0]:s[1] for s in sections}

    for idx, key in enumerate(cost_keys):
        if idx in sec_idx and sec_idx[idx]:
            ws2.cell(r2,1,f"  — {sec_idx[idx]}").font=FM; r2+=1
        lbl = COST_LABELS.get(key, key)
        cost_rows[key] = wr2(key, f"  {lbl}")

    # TOTAL COSTS (SUM formula)
    all_cost_row_nums = list(cost_rows.values())
    ws2.cell(r2,1,"TOTAL COSTS").font=FS; ws2.cell(r2,1).border=BB
    for m in range(1,49):
        c=col(m); formula="="+"+".join(f"{c}{rr}" for rr in all_cost_row_nums)
        ws2.cell(r2,mc(m),formula).font=FS; ws2.cell(r2,mc(m)).number_format=FC; ws2.cell(r2,mc(m)).border=BB
    for y in range(1,5):
        ycc=yc(y); formula="="+"+".join(f"{ycc}{rr}" for rr in all_cost_row_nums)
        ws2.cell(r2,M+1+y,formula).font=FS; ws2.cell(r2,M+1+y).number_format=FC
    r_tc=r2; r2+=2

    # PROFITABILITY
    ws2.cell(r2,1,"PROFITABILITY").font=FS; ws2.cell(r2,1).fill=PS; r2+=1
    ws2.cell(r2,1,"EBITDA").font=FS; ws2.cell(r2,1).border=BB
    for m in range(1,49):
        c=col(m); ws2.cell(r2,mc(m),f"={c}{r_rev}-{c}{r_tc}").font=FS; ws2.cell(r2,mc(m)).number_format=FC; ws2.cell(r2,mc(m)).border=BB
    for y in range(1,5):
        ycc=yc(y); ws2.cell(r2,M+1+y,f"={ycc}{r_rev}-{ycc}{r_tc}").font=FS; ws2.cell(r2,M+1+y).number_format=FC
    r_eb=r2; r2+=1

    ws2.cell(r2,1,"EBITDA margin").font=FN
    for m in range(1,49): c=col(m); ws2.cell(r2,mc(m),f"=IF({c}{r_rev}=0,0,{c}{r_eb}/{c}{r_rev})").font=FF; ws2.cell(r2,mc(m)).number_format=FP
    r2+=1

    ws2.cell(r2,1,"Cumulative EBITDA").font=FS
    for m in range(1,49):
        c=col(m); ws2.cell(r2,mc(m),f"={c}{r_eb}" if m==1 else f"={col(m-1)}{r2}+{c}{r_eb}").font=FS; ws2.cell(r2,mc(m)).number_format=FC
    for y in range(1,5): ec=col(y*12); ws2.cell(r2,M+1+y,f"={ec}{r2}").font=FS; ws2.cell(r2,M+1+y).number_format=FC
    r2+=1

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 3: CASH FLOW — mirrors P&L with cross-sheet formulas
    # ═══════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Cash Flow"); ws3.sheet_properties.tabColor = GOLD
    ws3.column_dimensions["A"].width = 44
    for m in range(1,49): ws3.column_dimensions[col(m)].width = 13
    for y in range(1,5): ws3.column_dimensions[yc(y)].width = 16
    ws3.freeze_panes = "B5"

    r3=1; ws3.cell(r3,1,"LUMINA — CASH FLOW (CHF) — ZUG/BAAR").font=FT; r3+=2
    for m in range(1,49): yr=(m-1)//12+1; ws3.cell(r3,mc(m),f"Y{yr}").font=Font(name="Calibri",size=9,color=TERRA); ws3.cell(r3,mc(m)).alignment=AC
    for y in range(1,5): ws3.cell(r3,M+1+y,f"Y{y}").font=FH; ws3.cell(r3,M+1+y).fill=PH; ws3.cell(r3,M+1+y).alignment=AC
    r3+=1
    ws3.cell(r3,1,"Month").font=FH; ws3.cell(r3,1).fill=PH
    for m in range(1,49): ws3.cell(r3,mc(m),f"M{m}").font=FH; ws3.cell(r3,mc(m)).fill=PH; ws3.cell(r3,mc(m)).alignment=AC
    r3+=1

    # INFLOWS
    ws3.cell(r3,1,"CASH INFLOWS").font=FS; ws3.cell(r3,1).fill=PS; r3+=1
    # Capital
    ws3.cell(r3,1,"  Initial capital (CHF 200K)").font=FF
    ws3.cell(r3,mc(1),CAPITAL).font=FF; ws3.cell(r3,mc(1)).number_format=FC
    for m in range(2,49): ws3.cell(r3,mc(m),0).font=FF; ws3.cell(r3,mc(m)).number_format=FC
    for y in range(1,5): ws3.cell(r3,M+1+y,CAPITAL if y==1 else 0).font=FF; ws3.cell(r3,M+1+y).number_format=FC
    r_cap=r3; r3+=1
    # Revenue (→ P&L)
    ws3.cell(r3,1,"  Revenue").font=FF
    for m in range(1,49): c=col(m); ws3.cell(r3,mc(m),f"='Monthly P&L'!{c}{r_rev}").font=FF; ws3.cell(r3,mc(m)).number_format=FC
    for y in range(1,5): ycc=yc(y); ws3.cell(r3,M+1+y,f"='Monthly P&L'!{ycc}{r_rev}").font=FF; ws3.cell(r3,M+1+y).number_format=FC
    r_crev=r3; r3+=1
    # Total inflows
    ws3.cell(r3,1,"TOTAL INFLOWS").font=FS; ws3.cell(r3,1).border=BB
    for m in range(1,49): c=col(m); ws3.cell(r3,mc(m),f"={c}{r_cap}+{c}{r_crev}").font=FS; ws3.cell(r3,mc(m)).number_format=FC; ws3.cell(r3,mc(m)).border=BB
    for y in range(1,5): ycc=yc(y); ws3.cell(r3,M+1+y,f"={ycc}{r_cap}+{ycc}{r_crev}").font=FS; ws3.cell(r3,M+1+y).number_format=FC
    r_in=r3; r3+=2

    # OUTFLOWS — every item from P&L
    ws3.cell(r3,1,"CASH OUTFLOWS — EVERY ITEM").font=FS; ws3.cell(r3,1).fill=PS; r3+=1
    cf_out_rows = []
    for idx, key in enumerate(cost_keys):
        if idx in sec_idx and sec_idx[idx]:
            ws3.cell(r3,1,f"  — {sec_idx[idx]}").font=FM; r3+=1
        lbl = COST_LABELS.get(key, key)
        pl_row = cost_rows[key]
        ws3.cell(r3,1,f"  {lbl}").font=FF
        for m in range(1,49): c=col(m); ws3.cell(r3,mc(m),f"='Monthly P&L'!{c}{pl_row}").font=FF; ws3.cell(r3,mc(m)).number_format=FC
        for y in range(1,5): ycc=yc(y); ws3.cell(r3,M+1+y,f"='Monthly P&L'!{ycc}{pl_row}").font=FF; ws3.cell(r3,M+1+y).number_format=FC
        cf_out_rows.append(r3); r3+=1

    # Total outflows
    ws3.cell(r3,1,"TOTAL OUTFLOWS").font=FS; ws3.cell(r3,1).border=BB
    for m in range(1,49):
        c=col(m); formula="="+"+".join(f"{c}{rr}" for rr in cf_out_rows)
        ws3.cell(r3,mc(m),formula).font=FS; ws3.cell(r3,mc(m)).number_format=FC; ws3.cell(r3,mc(m)).border=BB
    for y in range(1,5):
        ycc=yc(y); formula="="+"+".join(f"{ycc}{rr}" for rr in cf_out_rows)
        ws3.cell(r3,M+1+y,formula).font=FS; ws3.cell(r3,M+1+y).number_format=FC
    r_out=r3; r3+=2

    # NET
    ws3.cell(r3,1,"NET CASH FLOW").font=FS; ws3.cell(r3,1).fill=PS; r3+=1
    ws3.cell(r3,1,"Net cash flow").font=FS; ws3.cell(r3,1).border=BB
    for m in range(1,49): c=col(m); ws3.cell(r3,mc(m),f"={c}{r_in}-{c}{r_out}").font=FS; ws3.cell(r3,mc(m)).number_format=FC; ws3.cell(r3,mc(m)).border=BB
    for y in range(1,5): ycc=yc(y); ws3.cell(r3,M+1+y,f"={ycc}{r_in}-{ycc}{r_out}").font=FS; ws3.cell(r3,M+1+y).number_format=FC
    r_net=r3; r3+=1

    # Closing
    ws3.cell(r3,1,"Closing cash balance").font=FS; ws3.cell(r3,1).border=BB
    for m in range(1,49):
        c=col(m)
        ws3.cell(r3,mc(m),f"={c}{r_net}" if m==1 else f"={col(m-1)}{r3}+{c}{r_net}").font=FS
        ws3.cell(r3,mc(m)).number_format=FC; ws3.cell(r3,mc(m)).border=BB
    for y in range(1,5): ec=col(y*12); ws3.cell(r3,M+1+y,f"={ec}{r3}").font=FS; ws3.cell(r3,M+1+y).number_format=FC
    r_close=r3; r3+=1

    # Runway
    # Net burn = outflows - revenue. If net burn <= 0, company is profitable = infinite runway.
    ws3.cell(r3,1,"Net burn (costs - revenue)").font=FN
    for m in range(1,49):
        c=col(m)
        ws3.cell(r3,mc(m),f"={c}{r_out}-{c}{r_crev}").font=FF
        ws3.cell(r3,mc(m)).number_format=FC
    r_nb=r3; r3+=1

    ws3.cell(r3,1,"Runway (months, net burn)").font=FN
    for m in range(1,49):
        c=col(m)
        # If net burn <= 0 (profitable), runway = infinite (show 999)
        ws3.cell(r3,mc(m),f'=IF({c}{r_nb}<=0,999,{c}{r_close}/{c}{r_nb})').font=FF
        ws3.cell(r3,mc(m)).number_format='0.0"mo"'
    r3+=1

    ws3.cell(r3,1,"6-month runway clause met?").font=FM
    ws3.cell(r3+1,1,"(runway >= 6 OR profitable = always met)").font=FM
    for m in range(1,49):
        c=col(m)
        if m==1:
            ws3.cell(r3,mc(m),"YES").font=FF
        else:
            prev=col(m-1)
            # If prev month net burn <= 0 (profitable), clause auto-met
            # Otherwise check: prev_cash >= prev_net_burn × 6
            ws3.cell(r3,mc(m),f'=IF({prev}{r_nb}<=0,"YES",IF({prev}{r_close}>={prev}{r_nb}*6,"YES","NO"))').font=FF
    r3+=2

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 4: ZUG vs ZURICH COMPARISON
    # ═══════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Zug vs Zurich"); ws4.sheet_properties.tabColor = TERRA
    ws4.column_dimensions["A"].width = 48
    for ci in range(2,6): ws4.column_dimensions[get_column_letter(ci)].width = 20
    r4=1; ws4.cell(r4,1,"ZUG/BAAR vs ZURICH — FULL COMPARISON").font=FT; r4+=2

    # Tax comparison
    ws4.cell(r4,1,"CORPORATE PROFIT TAX").font=FS; ws4.cell(r4,1).fill=PS; r4+=1
    for ci,h in enumerate(["","Baar/Zug","Zurich City","Annual Savings"],1):
        ws4.cell(r4,ci,h).font=FH; ws4.cell(r4,ci).fill=PH; ws4.cell(r4,ci).alignment=AC
    r4+=1
    ws4.cell(r4,1,"Effective rate").font=FN; ws4.cell(r4,2,"11.82%").font=FI; ws4.cell(r4,3,"19.59%").font=FI; ws4.cell(r4,4,"-7.77pp").font=FS; r4+=1
    for profit in [100000,500000,1000000]:
        zug=round(profit*0.1182); zur=round(profit*0.1959); save=zur-zug
        ws4.cell(r4,1,f"Tax on CHF {profit:,} profit").font=FN
        ws4.cell(r4,2,zug).font=FF; ws4.cell(r4,2).number_format=FC
        ws4.cell(r4,3,zur).font=FF; ws4.cell(r4,3).number_format=FC
        ws4.cell(r4,4,save).font=FS; ws4.cell(r4,4).number_format=FC; r4+=1
    r4+=1

    # After-tax profit by year
    ws4.cell(r4,1,"AFTER-TAX PROFIT BY YEAR").font=FS; ws4.cell(r4,1).fill=PS; r4+=1
    for ci,h in enumerate(["Year","EBITDA","Tax (Zug)","After-tax (Zug)","Tax (Zurich)","After-tax (ZH)","Zug saves"],1):
        ws4.cell(r4,ci,h).font=FH; ws4.cell(r4,ci).fill=PH; ws4.cell(r4,ci).alignment=AC
    ws4.column_dimensions["E"].width=18; ws4.column_dimensions["F"].width=20; ws4.column_dimensions["G"].width=16
    r4+=1
    cumulative_save = 0
    for y in range(4):
        ebitda = sum(data[i]["ebitda"] for i in range(y*12,(y+1)*12))
        taxable = max(0, ebitda)  # losses carried forward (simplified)
        t_zug = round(taxable * 0.1182); t_zur = round(taxable * 0.1959)
        save = t_zur - t_zug; cumulative_save += save
        ws4.cell(r4,1,f"Y{y+1}").font=FN
        ws4.cell(r4,2,round(ebitda)).font=FF; ws4.cell(r4,2).number_format=FC
        ws4.cell(r4,3,t_zug).font=FF; ws4.cell(r4,3).number_format=FC
        ws4.cell(r4,4,round(ebitda-t_zug)).font=FF; ws4.cell(r4,4).number_format=FC
        ws4.cell(r4,5,t_zur).font=FF; ws4.cell(r4,5).number_format=FC
        ws4.cell(r4,6,round(ebitda-t_zur)).font=FF; ws4.cell(r4,6).number_format=FC
        ws4.cell(r4,7,save).font=FS; ws4.cell(r4,7).number_format=FC; r4+=1
    ws4.cell(r4,1,"CUMULATIVE TAX SAVINGS (ZUG)").font=FS
    ws4.cell(r4,7,cumulative_save).font=FS; ws4.cell(r4,7).number_format=FC; ws4.cell(r4,7).border=BB; r4+=2

    # Other differences
    ws4.cell(r4,1,"OTHER DIFFERENCES").font=FS; ws4.cell(r4,1).fill=PS; r4+=1
    diffs = [
        ("Incorporation notary fee","CHF 1,000","CHF 1,500","CHF 500 more"),
        ("FAK employer rate","1.35%","0.98%","~CHF 400/yr cheaper in ZH per employee"),
        ("Capital tax rate","0.07%","0.17%","CHF 200 more/yr at CHF 200K equity"),
        ("Treuhand costs","Similar","Similar","10-20% higher in ZH on average"),
        ("Domiciliation","CHF 29/mo basic","CHF 50/mo basic","Minimal difference"),
        ("Insurance","Same","Same","Not canton-dependent"),
    ]
    for item,zug,zur,note in diffs:
        ws4.cell(r4,1,item).font=FN; ws4.cell(r4,2,zug).font=FI; ws4.cell(r4,3,zur).font=FI; ws4.cell(r4,4,note).font=FM; r4+=1
    r4+=1
    ws4.cell(r4,1,"VERDICT: Zug/Baar saves CHF 38K-78K/year in corporate tax alone.").font=FS
    ws4.cell(r4+1,1,"Incorporation cost difference is negligible. Choose Zug.").font=FM

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 5: VALUATION & EXIT
    # ═══════════════════════════════════════════════════════════════════
    ws5 = wb.create_sheet("Valuation & Exit"); ws5.sheet_properties.tabColor = DEEP
    ws5.column_dimensions["A"].width = 44
    for ci in range(2,8): ws5.column_dimensions[get_column_letter(ci)].width = 18
    r5=1; ws5.cell(r5,1,"LUMINA — VALUATION & EXIT").font=FT; r5+=2
    ws5.cell(r5,1,"ARR MILESTONES").font=FS; ws5.cell(r5,1).fill=PS; r5+=1
    for ci,h in enumerate(["","ARR CHF","ARR EUR","Paying","EBITDA/mo","Cash"],1):
        ws5.cell(r5,ci,h).font=FH; ws5.cell(r5,ci).fill=PH; ws5.cell(r5,ci).alignment=AC
    r5+=1
    KICKER=[(0,.20),(30e6,.23),(50e6,.25),(100e6,.30)]
    for mm in [12,24,36,48]:
        d=data[mm-1]; ws5.cell(r5,1,f"M{mm}").font=FN
        ws5.cell(r5,2,d["arr"]).font=FF; ws5.cell(r5,2).number_format=FC
        ws5.cell(r5,3,round(d["arr_eur"])).font=FF; ws5.cell(r5,3).number_format=FE
        ws5.cell(r5,4,d["total_pay"]).font=FF
        ws5.cell(r5,5,d["ebitda"]).font=FF; ws5.cell(r5,5).number_format=FC
        ws5.cell(r5,6,round(d["closing"])).font=FF; ws5.cell(r5,6).number_format=FC; r5+=1
    r5+=1
    ws5.cell(r5,1,"VALUATION MATRIX (CHF)").font=FS; ws5.cell(r5,1).fill=PS; r5+=1
    for ci,x in enumerate([6,8,10,12],2): ws5.cell(r5,ci,f"{x}x").font=FH; ws5.cell(r5,ci).fill=PH; ws5.cell(r5,ci).alignment=AC
    r5+=1
    for mm in [12,24,36,48]:
        d=data[mm-1]; ws5.cell(r5,1,f"M{mm}").font=FN
        for ci,x in enumerate([6,8,10,12],2): ws5.cell(r5,ci,round(d["arr"]*x)).font=FF; ws5.cell(r5,ci).number_format=FC
        r5+=1
    r5+=1
    ws5.cell(r5,1,"MARCO'S EQUITY VALUE (10x exit)").font=FS; ws5.cell(r5,1).fill=PS; r5+=1
    for mm in [12,24,36,48]:
        d=data[mm-1]; cv=d["arr"]*10; cv_eur=c2e(cv)
        kp=0.20
        for th,pct in KICKER:
            if cv_eur>=th: kp=pct
        ws5.cell(r5,1,f"M{mm}: {kp*100:.0f}% kicker").font=FN
        ws5.cell(r5,2,round(cv*kp)).font=FF; ws5.cell(r5,2).number_format=FC
        ws5.cell(r5,3,f"on CHF {cv:,.0f} valuation").font=FM; r5+=1

    # ═══════════════════════════════════════════════════════════════════
    # SHEET 6: SENSITIVITY
    # ═══════════════════════════════════════════════════════════════════
    ws6 = wb.create_sheet("Sensitivity"); ws6.sheet_properties.tabColor = GOLD
    ws6.column_dimensions["A"].width = 52
    for ci in range(2,6): ws6.column_dimensions[get_column_letter(ci)].width = 18
    r6=1; ws6.cell(r6,1,"LUMINA — SENSITIVITY & RUNWAY").font=FT; r6+=2
    be = next((i+1 for i,d in enumerate(data) if d["ebitda"]>0), None)
    mc_d = min(data, key=lambda d: d["closing"])
    ws6.cell(r6,1,"BREAKEVEN").font=FS; ws6.cell(r6,1).fill=PS; r6+=1
    ws6.cell(r6,1,"First EBITDA-positive").font=FN; ws6.cell(r6,2,f"M{be}" if be else "N/A").font=FS; r6+=1
    ws6.cell(r6,1,"Lowest cash").font=FN; ws6.cell(r6,2,round(mc_d["closing"])).font=FF; ws6.cell(r6,2).number_format=FC; ws6.cell(r6,3,f"M{data.index(mc_d)+1}").font=FM; r6+=1
    ws6.cell(r6,1,"Max drawdown").font=FN; ws6.cell(r6,2,round(CAPITAL-mc_d["closing"])).font=FF; ws6.cell(r6,2).number_format=FC; r6+=2
    ws6.cell(r6,1,"CEO STEP-UP TIMELINE").font=FS; ws6.cell(r6,1).fill=PS; r6+=1
    seen=set()
    for i,d in enumerate(data):
        if d["ceo_sal"] not in seen:
            seen.add(d["ceo_sal"])
            ws6.cell(r6,1,f"M{i+1}: CHF {d['ceo_sal']:,}/yr → CHF {d['ceo_total']:,}/mo").font=FN; r6+=1
    r6+=1
    # VESTING
    ws6.cell(r6,1,"VESTING STATUS").font=FS; ws6.cell(r6,1).fill=PS; r6+=1
    VEST=[(12,.05,"15+ paying"),(24,.07,"ARR≥€500K"),(36,.08,"ARR≥€1.5M")]
    cum=.20
    for mv,pct,cond in VEST:
        cum+=pct; d=data[mv-1]
        met=(d["total_pay"]>=15) if mv==12 else (d["arr_eur"]>=500000) if mv==24 else (d["arr_eur"]>=1500000)
        ws6.cell(r6,1,f"M{mv}: +{pct*100:.0f}% if {cond}").font=FN
        ws6.cell(r6,2,"MET" if met else "NOT MET").font=Font(name="Calibri",size=10,bold=True,color="006600" if met else "CC0000")
        ws6.cell(r6,3,f"Cumul {cum*100:.0f}%").font=FM; r6+=1

    # ═══════════════════════════════════════════════════════════════════
    # SAVE & VALIDATE
    # ═══════════════════════════════════════════════════════════════════
    wb.save(OUTPUT)
    print(f"\nSaved {OUTPUT}")
    print("="*90)

    d12=data[11]; d48=data[47]
    be_m=next((i+1 for i,d in enumerate(data) if d["ebitda"]>0),None)
    mc_d=min(data,key=lambda d:d["closing"]); mc_m=data.index(mc_d)+1

    print(f"FX: €1K=CHF {e2c(1000)} | $600=CHF {u2c(600)}")
    checks=[("Breakeven<=M12",be_m and be_m<=12,f"M{be_m}"),
            ("Cash>0",mc_d["closing"]>0,f"CHF {mc_d['closing']:,.0f} at M{mc_m}"),
            ("M12 paying>=30",d12["total_pay"]>=30,str(d12["total_pay"])),
            ("M12 ARR~€540K",abs(d12["arr_eur"]-540000)<50000,f"€{d12['arr_eur']:,.0f}"),
            ("Capital=200K",CAPITAL==200000,"OK")]
    for n,ok,det in checks: print(f"[{'PASS' if ok else 'FAIL'}] {n}: {det}")

    print(f"\nP&L has {len(cost_keys)} individual cost rows + {len(CHAINS)} chain rows")
    print(f"Cash Flow has {len(cf_out_rows)} outflow rows with ='Monthly P&L'! formulas")

    print(f"\nM1 EVERY COST:")
    for key in cost_keys:
        v=data[0][key]
        if v>0: print(f"  {COST_LABELS.get(key,key):50s} CHF {v:>6,}")
    print(f"  {'TOTAL':50s} CHF {data[0]['total_costs']:>6,}")

    print(f"\nKEY METRICS:")
    for mm in [12,24,36,48]:
        d=data[mm-1]; print(f"  M{mm}: {d['total_pay']:3d} pay | rev CHF {d['rev']:>8,} | ARR €{d['arr_eur']:>10,.0f} | EBITDA CHF {d['ebitda']:>8,} | cash CHF {d['closing']:>10,.0f}")

    print(f"\nANNUAL:")
    for y in range(4):
        rv=sum(data[i]["rev"] for i in range(y*12,(y+1)*12))
        co=sum(data[i]["total_costs"] for i in range(y*12,(y+1)*12))
        print(f"  Y{y+1}: Rev CHF {rv:>10,} | Cost CHF {co:>10,} | EBITDA CHF {rv-co:>10,}")

    print(f"\nZUG vs ZURICH TAX (cumulative 4yr):")
    tot_ebitda = sum(max(0,sum(data[i]["ebitda"] for i in range(y*12,(y+1)*12))) for y in range(4))
    print(f"  Zug tax: CHF {round(tot_ebitda*0.1182):,}")
    print(f"  Zurich tax: CHF {round(tot_ebitda*0.1959):,}")
    print(f"  4-year savings: CHF {round(tot_ebitda*(0.1959-0.1182)):,}")

if __name__ == "__main__":
    build()
